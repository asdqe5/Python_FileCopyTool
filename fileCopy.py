# _*_ coding: utf-8 _*_

# Comp Publish Tool
#
# Description : Copy 버튼을 눌렀을 때, 테이블에 있는 샷코드를 기준으로 해당 경로로 복사하는 함수

import os
import sys
import re
import openpyxl
import math

import shutil
from distutils.dir_util import copy_tree

from PySide2 import QtWidgets, QtGui, QtCore

SHOW_PATH = os.getenv("SHOW_PATH")
TD_PATH = os.getenv("TD_PATH")

if "{}/api" not in sys.path:
    sys.path.append('{}/api'.format(TD_PATH))

import shotgun_api3

sg = shotgun_api3.Shotgun("https://road101.shotgunstudio.com", script_name="authentication_script", api_key="vZnk#jbmopl6oqispvodazfyn")

def palteCopyFunc(self):
    '''
    테이블의 샷코드를 기준으로 
    플레이트 및 소스 파일을 복사하는 함수
    '''
    # 프로젝트 콤보박스 확인
    projectName = self.ui.plate_project_comboBox.currentText()
    if projectName == "":
        dial = QtWidgets.QMessageBox()
        dial.setWindowTitle("Error")
        dial.setText(u"프로젝트를 선택해주세요.")
        ok_btn = dial.addButton("OK", QtWidgets.QMessageBox.YesRole)
        dial.exec_()
        return
    
    # 복사할 경로가 정해졌는지 확인
    pathToCopy = self.ui.plate_pathtocopy_lineEdit.text()
    if pathToCopy == "":
        dial = QtWidgets.QMessageBox()
        dial.setWindowTitle("Error")
        dial.setText(u"복사할 경로를 지정해주세요.")
        ok_btn = dial.addButton("OK", QtWidgets.QMessageBox.YesRole)
        dial.exec_()
        return
    
    # 엑셀 라디오 버튼이 체크되어있다면 엑셀 파일 경로가 정해졌는지 확인
    if self.ui.plate_excel_radioButton.isChecked():
        excelPath = self.ui.plate_excelFile_lineEdit.text()
        if excelPath == "":
            dial = QtWidgets.QMessageBox()
            dial.setWindowTitle("Error")
            dial.setText(u"엑셀 파일을 먼저 선택해주세요.")
            ok_btn = dial.addButton("OK", QtWidgets.QMessageBox.YesRole)
            dial.exec_()
            return
    
    # 샷건 라디오 버튼이 체크되어있다면 벤더 콤보 박스가 체크되었는지 확인
    if self.ui.plate_shotgun_radioButton.isChecked():
        vendor = self.ui.plate_vendor_comboBox.currentText()
        if vendor == "":
            dial = QtWidgets.QMessageBox()
            dial.setWindowTitle("Error")
            dial.setText(u"벤더를 선택해주세요.")
            ok_btn = dial.addButton("OK", QtWidgets.QMessageBox.YesRole)
            dial.exec_()
            return

    # 샷코드를 읽어왔는지 확인
    tableRow = self.ui.plate_fileCopy_tableWidget.rowCount()
    if tableRow == 0:
        dial = QtWidgets.QMessageBox()
        dial.setWindowTitle("Error")
        dial.setText("Copy할 데이터가 존재하지 않습니다.")
        ok_btn = dial.addButton("OK", QtWidgets.QMessageBox.YesRole)
        dial.exec_()
        return

    shotCodeList = []

    # 툴 메뉴 비활성화
    self.ui.plate_project_comboBox.setEnabled(False)
    self.ui.plate_excel_radioButton.setEnabled(False)
    self.ui.plate_shotgun_radioButton.setEnabled(False)
    self.ui.plate_pathToCopy_pushButton.setEnabled(False)
    self.ui.plate_vendor_comboBox.setEnabled(False)
    self.ui.plate_vendorLoad_pushButton.setEnabled(False)
    self.ui.plate_excelFile_pushButton.setEnabled(False)
    self.ui.plate_excelLoad_pushButton.setEnabled(False)
    self.ui.plate_copy_pushButton.setEnabled(False)
    self.ui.plate_export_pushButton.setEnabled(False)
    self.ui.plate_cancel_pushButton.setEnabled(False)

    for r in range(tableRow):

        copied = False
        shotCode = self.ui.plate_fileCopy_tableWidget.item(r, self.plate_column_idx_lookup["ShotCode"]).text()
        shotPath = os.path.join(SHOW_PATH, projectName, "shot", shotCode.split("_")[0], shotCode.split("_")[1])

        # 샷코드가 규칙에 맞는지 확인
        if not re.match("s[0-9]+_c[0-9]+", shotCode):
            if not re.match("s[0-9]+_c[a-z]+", shotCode):
                item = QtWidgets.QTableWidgetItem("샷코드가 규칙에 맞지 않습니다.")
                item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
                self.ui.plate_fileCopy_tableWidget.setItem(r, self.plate_column_idx_lookup["Error"], item)

                color = QtGui.QColor(255, 0, 0, 125)
                self.ui.plate_fileCopy_tableWidget.item(r, self.plate_column_idx_lookup["ShotCode"]).setBackground(color)
                QtCore.QCoreApplication.processEvents()
                continue
        
        # 동일한 샷코드가 있는지 확인
        if shotCode in shotCodeList:
            item = QtWidgets.QTableWidgetItem("이미 존재하는 샷코드입니다.")
            item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
            self.ui.plate_fileCopy_tableWidget.setItem(r, self.plate_column_idx_lookup["Error"], item)

            color = QtGui.QColor(255, 0, 0, 125)
            self.ui.plate_fileCopy_tableWidget.item(r, self.plate_column_idx_lookup["ShotCode"]).setBackground(color)
            QtCore.QCoreApplication.processEvents()
            continue
        shotCodeList.append(shotCode)
        
        # 해당 샷코드에 맞는 폴더 경로가 존재하는지 확인한다.
        if not os.path.exists(shotPath):
            item = QtWidgets.QTableWidgetItem("경로가 존재하지 않습니다.")
            item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
            self.ui.plate_fileCopy_tableWidget.setItem(r, self.plate_column_idx_lookup["Error"], item)

            color = QtGui.QColor(255, 0, 0, 125)
            self.ui.plate_fileCopy_tableWidget.item(r, self.plate_column_idx_lookup["ShotCode"]).setBackground(color)
            QtCore.QCoreApplication.processEvents()
            continue

        # plate 폴더에서 최신 버전이 존재한다면 복사한다. -> 플레이트와 리타임 플레이트 확인
        platePath = os.path.join(shotPath, "plate")
        if os.path.exists(platePath):
            plate = os.listdir(platePath)
            plate.sort()

            plateList = []
            rplateList = []
            for p in plate:
                # 숨겨져 있는지 확인
                if p.startswith("."):
                    continue

                # 폴더인지 확인
                if not os.path.isdir(os.path.join(platePath, p)):
                    continue

                # 플레이트인지 리타임 플레이트인지 구분
                if re.match("plate_v0[0-9]+", p):
                    plateList.append(p)

                if re.match("plate_v[1-9][0-9]+", p):
                    rplateList.append(p)
            
            # 리타임 플레이트 리스트가 있다면 최신 버전을 같은 경로로 복사한다.
            if rplateList != []:
                copied = True
                maxValue = int(rplateList[0].split("v")[1])
                lastVersion = rplateList[0]
                for rp in rplateList:
                    if maxValue < int(rp.split("v")[1]):
                        maxValue = int(rp.split("v")[1])
                        lastVersion = rp

                # 디렉토리 경로가 없다면 경로에 맞게 디렉토리를 생성해준다
                pathToCopyRPlate = os.path.join(pathToCopy, shotCode.split("_")[0], shotCode.split("_")[1], "plate", lastVersion)
                if not os.path.isdir(pathToCopyRPlate): # 복사하려는 플레이트 폴더가 존재하지 않을 때
                    os.makedirs(pathToCopyRPlate)

                    # 복사하려는 플레이트 버전보다 낮은 버전이 있는지 확인
                    copiedRPlate = os.listdir(os.path.dirname(pathToCopyRPlate))
                    
                    # 버전이 낮다면 삭제한다.
                    for crp in copiedRPlate:
                        version = int(crp.split("v")[1])
                        if version < maxValue:
                            shutil.rmtree(os.path.join(os.path.dirname(pathToCopyRPlate), crp))

                    # 해당 경로로 최신 버전의 플레이트 폴더를 복사한다
                    copyRPlatePath = os.path.join(platePath, lastVersion)
                    copy_tree(copyRPlatePath, pathToCopyRPlate)

                # 복사된 타입과 버전을 테이블에 적어준다.
                item = QtWidgets.QTableWidgetItem(lastVersion)
                item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
                self.ui.plate_fileCopy_tableWidget.setItem(r, self.plate_column_idx_lookup["r_Plate"], item)
                QtCore.QCoreApplication.processEvents()

            # 플레이트 리스트가 있고 리타임 플레이트가 없다면 최신 버전을 같은 경로로 복사한다.
            elif plateList != []:
                copied = True
                maxValue = int(plateList[0].split("v")[1])
                lastVersion = plateList[0]
                for p in plateList:
                    if maxValue < int(p.split("v")[1]):
                        maxValue = int(p.split("v")[1])
                        lastVersion = p

                # 디렉토리 경로가 없다면 경로에 맞게 디렉토리를 생성해준다
                pathToCopyPlate = os.path.join(pathToCopy, shotCode.split("_")[0], shotCode.split("_")[1], "plate", lastVersion)
                if not os.path.isdir(pathToCopyPlate):
                    os.makedirs(pathToCopyPlate)
                    
                    # 복사하려는 플레이트 버전보다 낮은 버전이 있는지 확인
                    copiedPlate = os.listdir(os.path.dirname(pathToCopyPlate))

                    # 버전이 낮다면 삭제한다.
                    for cp in copiedPlate:
                        version = int(cp.split("v")[1])
                        if version < maxValue:
                            shutil.rmtree(os.path.join(os.path.dirname(pathToCopyPlate), cp))

                    # 해당 경로로 최신 버전의 플레이트 폴더를 복사한다
                    copyPlatePath = os.path.join(platePath, lastVersion)
                    copy_tree(copyPlatePath, pathToCopyPlate)

                # 복사된 타입과 버전을 테이블에 적어준다.
                item = QtWidgets.QTableWidgetItem(lastVersion)
                item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
                self.ui.plate_fileCopy_tableWidget.setItem(r, self.plate_column_idx_lookup["Plate"], item)
                QtCore.QCoreApplication.processEvents()

        # src 폴더에서 최신 버전이 존재한다면 복사한다. -> p_src 와 src 확인
        srcPath = os.path.join(shotPath, "src")
        if os.path.exists(srcPath):
            # src 폴더 내의 plate 폴더의 p_src 확인
            psrcPath = os.path.join(srcPath, "plate")
            if os.path.exists(psrcPath):
                psrc = os.listdir(psrcPath)
                psrc.sort()

                psrcDict = {}
                for ps in psrc:
                    # 숨겨져 있는지 확인
                    if ps.startswith("."):
                        continue

                    # 폴더인지 확인
                    if not os.path.isdir(os.path.join(psrcPath, ps)):
                        continue

                    # p_src 종류 확인
                    psrcType = re.findall("(p_src_[0-9]+)_v[0-9]+", ps)
                    if not psrcType[0] in psrcDict:
                        psrcDict[psrcType[0]] = [ps]
                    else:
                        psrcDict[psrcType[0]].append(ps)
                
                sortedPsrcDict = sorted(psrcDict.items())
                for value in sortedPsrcDict:
                    copied = True
                    psrcList = value[1]

                    maxValue = int(psrcList[0].split("v")[1])
                    lastVersion = psrcList[0]
                    for psr in psrcList:
                        if maxValue < int(psr.split("v")[1]):
                            maxValue = int(psr.split("v")[1])
                            lastVersion = psr
                    
                    # 디렉토리 경로가 없다면 경로게 맞게 디렉토리를 생성해준다
                    pathToCopyPSrc = os.path.join(pathToCopy, shotCode.split("_")[0], shotCode.split("_")[1], "src", "plate", lastVersion)
                    if not os.path.isdir(pathToCopyPSrc):
                        os.makedirs(pathToCopyPSrc)

                        copiedPSrc = os.listdir(os.path.dirname(pathToCopyPSrc))
                        
                        for cps in copiedPSrc:
                            cpsType = re.findall("(p_src_[0-9]+)_v[0-9]+", cps)
                            if cpsType[0] == value[0]:
                                if int(cps.split("v")[1]) < maxValue:
                                    shutil.rmtree(os.path.join(os.path.dirname(pathToCopyPSrc), cps))
                    
                        # 해당 경로로 최신 버전의 p_src 폴더를 복사한다.
                        copyPSrcPath = os.path.join(psrcPath, lastVersion)
                        copy_tree(copyPSrcPath, pathToCopyPSrc)

                    # 복사된 타입과 버전을 테이블에 적어준다.
                    copiedVersion = self.ui.plate_fileCopy_tableWidget.item(r, self.plate_column_idx_lookup["p_src"])
                    if copiedVersion:
                        item = QtWidgets.QTableWidgetItem(copiedVersion.text() + "\n" + lastVersion)
                    else:
                        item = QtWidgets.QTableWidgetItem(lastVersion)
                    item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
                    self.ui.plate_fileCopy_tableWidget.setItem(r, self.plate_column_idx_lookup["p_src"], item)
                    QtCore.QCoreApplication.processEvents()

            # src 폴더 내의 src 폴더의 src 확인
            ssrcPath = os.path.join(srcPath, "src")
            if os.path.exists(ssrcPath):
                ssrc = os.listdir(ssrcPath)
                ssrc.sort()

                ssrcDict = {}
                for ss in ssrc:
                    # 숨겨져 있는지 확인
                    if ss.startswith("."):
                        continue

                    # 폴더인지 확인
                    if not os.path.isdir(os.path.join(ssrcPath, ss)):
                        continue

                    # src 종류 확인
                    ssrcType = re.findall("(src_[0-9]+)_v[0-9]+", ss)
                    if not ssrcType[0] in ssrcDict:
                        ssrcDict[ssrcType[0]] = [ss]
                    else:
                        ssrcDict[ssrcType[0]].append(ss)

                sortedSsrcDict = sorted(ssrcDict.items())
                for value in sortedSsrcDict:
                    copied = True
                    ssrcList = value[1]

                    maxValue = int(ssrcList[0].split("v")[1])
                    lastVersion = ssrcList[0]
                    for ssr in ssrcList:
                        if maxValue < int(ssr.split("v")[1]):
                            maxValue = int(ssr.split("v")[1])
                            lastVersion = ssr

                    # 디렉토리 경로가 없다면 경로게 맞게 디렉토리를 생성해준다
                    pathToCopySSrc = os.path.join(pathToCopy, shotCode.split("_")[0], shotCode.split("_")[1], "src", "src", lastVersion)
                    if not os.path.isdir(pathToCopySSrc):
                        os.makedirs(pathToCopySSrc)
                        
                        copiedSSrc = os.listdir(os.path.dirname(pathToCopySSrc))

                        for css in copiedSSrc:
                            cssType = re.findall("(src_[0-9]+)_v[0-9]+", css)
                            if cssType[0] == value[0]:
                                if int(css.split("v")[1]) < maxValue:
                                    shutil.rmtree(os.path.join(os.path.dirname(pathToCopySSrc), css))
                    
                        # 해당 경로로 최신 버전의 src 폴더를 복사한다.
                        copySSrcPath = os.path.join(ssrcPath, lastVersion)
                        copy_tree(copySSrcPath, pathToCopySSrc)

                    # 복사된 타입과 버전을 테이블에 적어준다.
                    copiedVersion = self.ui.plate_fileCopy_tableWidget.item(r, self.plate_column_idx_lookup["src"])
                    if copiedVersion:
                        item = QtWidgets.QTableWidgetItem(copiedVersion.text() + "\n" + lastVersion)
                    else:
                        item = QtWidgets.QTableWidgetItem(lastVersion)
                    item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
                    self.ui.plate_fileCopy_tableWidget.setItem(r, self.plate_column_idx_lookup["src"], item)
                    QtCore.QCoreApplication.processEvents()
        
        if not copied:
            item = QtWidgets.QTableWidgetItem("복사할 파일이 존재하지 않습니다.")
            item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
            self.ui.plate_fileCopy_tableWidget.setItem(r, self.plate_column_idx_lookup["Error"], item)

            color = QtGui.QColor(255, 0, 0, 125)
            self.ui.plate_fileCopy_tableWidget.item(r, self.plate_column_idx_lookup["ShotCode"]).setBackground(color)
        else:
            color = QtGui.QColor(0, 255, 0, 125)
            self.ui.plate_fileCopy_tableWidget.item(r, self.plate_column_idx_lookup["ShotCode"]).setBackground(color)
        
        QtCore.QCoreApplication.processEvents()

    self.ui.plate_fileCopy_tableWidget.resizeRowsToContents()
    self.ui.plate_fileCopy_tableWidget.resizeColumnsToContents()

    # 버튼을 또 누르는 것을 방지하기 위해서 복사가 완료되면 버튼을 비활성화한다,
    self.ui.plate_copy_pushButton.setEnabled(False)

    # 툴 메뉴 활성화
    self.ui.plate_project_comboBox.setEnabled(True)
    self.ui.plate_excel_radioButton.setEnabled(True)
    self.ui.plate_shotgun_radioButton.setEnabled(True)
    self.ui.plate_pathToCopy_pushButton.setEnabled(True)
    if self.ui.plate_shotgun_radioButton.isChecked():
        self.ui.plate_vendor_comboBox.setEnabled(True)
        self.ui.plate_vendorLoad_pushButton.setEnabled(True)
    if self.ui.plate_excel_radioButton.isChecked():
        self.ui.plate_excelFile_pushButton.setEnabled(True)
        self.ui.plate_excelLoad_pushButton.setEnabled(True)
    self.ui.plate_export_pushButton.setEnabled(True)
    self.ui.plate_cancel_pushButton.setEnabled(True)

    dial = QtWidgets.QMessageBox()
    dial.setWindowTitle("Message")
    dial.setText("플레이트 복사가 모두 완료되었습니다.")
    ok_btn = dial.addButton("OK", QtWidgets.QMessageBox.YesRole)
    dial.exec_()

def exportPlateExcelFunc(self):
    '''
    테이블의 내용을 엑셀 파일로 
    정리하여 추출하는 함수이다.
    '''
    # 프로젝트 콤보박스 확인
    projectName = self.ui.plate_project_comboBox.currentText()
    if projectName == "":
        dial = QtWidgets.QMessageBox()
        dial.setWindowTitle("Error")
        dial.setText(u"프로젝트를 선택해주세요.")
        ok_btn = dial.addButton("OK", QtWidgets.QMessageBox.YesRole)
        dial.exec_()
        return
    
    # 복사할 경로가 정해졌는지 확인
    pathToCopy = self.ui.plate_pathtocopy_lineEdit.text()
    if pathToCopy == "":
        dial = QtWidgets.QMessageBox()
        dial.setWindowTitle("Error")
        dial.setText(u"복사할 경로를 지정해주세요.")
        ok_btn = dial.addButton("OK", QtWidgets.QMessageBox.YesRole)
        dial.exec_()
        return
    
    # 엑셀 라디오 버튼이 체크되어있으면 엑셀 파일 경로가 정해졌는지 확인
    if self.ui.plate_excel_radioButton.isChecked():
        excelPath = self.ui.plate_excelFile_lineEdit.text()
        if excelPath == "":
            dial = QtWidgets.QMessageBox()
            dial.setWindowTitle("Error")
            dial.setText(u"엑셀 파일을 먼저 선택해주세요.")
            ok_btn = dial.addButton("OK", QtWidgets.QMessageBox.YesRole)
            dial.exec_()
            return
    
    # 벤더 라디오 버튼이 체크되어있으면 벤더 콤보박스가 체크되었는지 확인
    if self.ui.plate_shotgun_radioButton.isChecked():
        vendor = self.ui.plate_vendor_comboBox.currentText()
        if vendor == "":
            dial = QtWidgets.QMessageBox()
            dial.setWindowTitle("Error")
            dial.setText(u"벤더를 먼저 선택해주세요.")
            ok_btn = dial.addButton("OK", QtWidgets.QMessageBox.YesRole)
            dial.exec_()
            return

    # 샷코드를 읽어왔는지 확인
    tableRow = self.ui.plate_fileCopy_tableWidget.rowCount()
    if tableRow == 0:
        dial = QtWidgets.QMessageBox()
        dial.setWindowTitle("Error")
        dial.setText("Copy할 데이터가 존재하지 않습니다.")
        ok_btn = dial.addButton("OK", QtWidgets.QMessageBox.YesRole)
        dial.exec_()
        return
    
    # 엑셀 만들기
    excelFileName = "{}_plate_copy.xlsx".format(projectName)
    excelFilePath = os.path.join(pathToCopy, excelFileName)

    write_wb = openpyxl.Workbook()
    write_ws = write_wb.create_sheet('plate_copy')
    write_ws = write_wb.active

    # 엑셀 제목
    write_ws.cell(1, 1, "ShotCode")
    write_ws.cell(1, 2, "Plate")
    write_ws.cell(1, 3, "r_Plate")
    write_ws.cell(1, 4, "p_src")
    write_ws.cell(1, 5, "src")
    write_ws.cell(1, 6, "Error")

    # 엑셀 내용 입력
    for r in range(tableRow):
        shotCode = self.ui.plate_fileCopy_tableWidget.item(r, self.plate_column_idx_lookup["ShotCode"]).text()
        write_ws.cell(r+2, 1, shotCode)

        if self.ui.plate_fileCopy_tableWidget.item(r, self.plate_column_idx_lookup["Plate"]):
            plate = self.ui.plate_fileCopy_tableWidget.item(r, self.plate_column_idx_lookup["Plate"]).text()
            write_ws.cell(r+2, 2, plate)

        if self.ui.plate_fileCopy_tableWidget.item(r, self.plate_column_idx_lookup["r_Plate"]):
            rPlate = self.ui.plate_fileCopy_tableWidget.item(r, self.plate_column_idx_lookup["r_Plate"]).text()
            write_ws.cell(r+2, 3, rPlate)
        
        if self.ui.plate_fileCopy_tableWidget.item(r, self.plate_column_idx_lookup["p_src"]):
            psrc = self.ui.plate_fileCopy_tableWidget.item(r, self.plate_column_idx_lookup["p_src"]).text()
            write_ws.cell(r+2, 4, psrc)
        
        if self.ui.plate_fileCopy_tableWidget.item(r, self.plate_column_idx_lookup["src"]):
            src = self.ui.plate_fileCopy_tableWidget.item(r, self.plate_column_idx_lookup["src"]).text()
            write_ws.cell(r+2, 5, src)
        
        if self.ui.plate_fileCopy_tableWidget.item(r, self.plate_column_idx_lookup["Error"]):
            error = self.ui.plate_fileCopy_tableWidget.item(r, self.plate_column_idx_lookup["Error"]).text()
            write_ws.cell(r+2, 6, error)

    # 엑셀 파일 셀 너비 및 높이 지정
    for col in write_ws.columns:
        max_width = 0
        col_name = col[0].column_letter
        for cell in col:
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            try:
                if '\n' in cell.value:
                    cell_value_list = cell.value.split("\n")
                    for cell_value in cell_value_list:
                        if len(str(cell_value.encode('utf-8'))) > max_width:
                            max_width = len(str(cell_value.encode('utf-8')))
                else:
                    if len(str(cell.value.encode('utf-8'))) > max_width:
                        max_width = len(str(cell.value.encode('utf-8')))
            except:
                continue
        adjusted_width = (max_width + 2) * 1.2
        write_ws.column_dimensions[col_name].width = adjusted_width
    
    for i, row in enumerate(write_ws):
        default_height = 15
        max_height = 0
        for cell in row:
            try:
                if '\n' in cell.value:
                    cell_value_list = cell.value.split("\n")
                    height = len(cell_value_list) * default_height
                    if max_height < height:
                        max_height = height
                else:
                    if max_height < default_height:
                        max_height = default_height
            except:
                continue
        write_ws.row_dimensions[i+1].height = max_height

    # 엑셀 파일 저장
    write_wb.save(excelFilePath)

    dial = QtWidgets.QMessageBox()
    dial.setWindowTitle("Message")
    dial.setText("해당 경로로 엑셀 파일이 저장되었습니다.")
    ok_btn = dial.addButton("OK", QtWidgets.QMessageBox.YesRole)
    dial.exec_()

def editCopyFunc(self):
    '''
    테이블의 샷코드를 기준으로 
    편집본을 복사하는 함수
    '''
    # 프로젝트 콤보박스 확인
    projectName = self.ui.edit_project_comboBox.currentText()
    if projectName == "":
        dial = QtWidgets.QMessageBox()
        dial.setWindowTitle("Error")
        dial.setText(u"프로젝트를 선택해주세요.")
        ok_btn = dial.addButton("OK", QtWidgets.QMessageBox.YesRole)
        dial.exec_()
        return
    
    # 복사할 경로가 정해졌는지 확인
    pathToCopy = self.ui.edit_pathtocopy_lineEdit.text()
    if pathToCopy == "":
        dial = QtWidgets.QMessageBox()
        dial.setWindowTitle("Error")
        dial.setText(u"복사할 경로를 지정해주세요.")
        ok_btn = dial.addButton("OK", QtWidgets.QMessageBox.YesRole)
        dial.exec_()
        return
    
    # 엑셀 라디오 버튼이 체크되어있다면 엑셀 파일 경로가 정해졌는지 확인
    if self.ui.edit_excel_radioButton.isChecked():
        excelPath = self.ui.edit_excelFile_lineEdit.text()
        if excelPath == "":
            dial = QtWidgets.QMessageBox()
            dial.setWindowTitle("Error")
            dial.setText(u"엑셀 파일을 먼저 선택해주세요.")
            ok_btn = dial.addButton("OK", QtWidgets.QMessageBox.YesRole)
            dial.exec_()
            return
    
    # 샷건 라디오 버튼이 체크되어있다면 벤더 콤보 박스가 체크되었는지 확인
    if self.ui.edit_shotgun_radioButton.isChecked():
        vendor = self.ui.edit_vendor_comboBox.currentText()
        if vendor == "":
            dial = QtWidgets.QMessageBox()
            dial.setWindowTitle("Error")
            dial.setText(u"벤더를 선택해주세요.")
            ok_btn = dial.addButton("OK", QtWidgets.QMessageBox.YesRole)
            dial.exec_()
            return

    # 샷코드를 읽어왔는지 확인
    tableRow = self.ui.edit_fileCopy_tableWidget.rowCount()
    if tableRow == 0:
        dial = QtWidgets.QMessageBox()
        dial.setWindowTitle("Error")
        dial.setText("Copy할 데이터가 존재하지 않습니다.")
        ok_btn = dial.addButton("OK", QtWidgets.QMessageBox.YesRole)
        dial.exec_()
        return

    shotCodeList = []

    # 툴 메뉴 비활성화
    self.ui.edit_project_comboBox.setEnabled(False)
    self.ui.edit_excel_radioButton.setEnabled(False)
    self.ui.edit_shotgun_radioButton.setEnabled(False)
    self.ui.edit_pathToCopy_pushButton.setEnabled(False)
    self.ui.edit_vendor_comboBox.setEnabled(False)
    self.ui.edit_vendorLoad_pushButton.setEnabled(False)
    self.ui.edit_excelFile_pushButton.setEnabled(False)
    self.ui.edit_excelLoad_pushButton.setEnabled(False)
    self.ui.edit_copy_pushButton.setEnabled(False)
    self.ui.edit_export_pushButton.setEnabled(False)
    self.ui.edit_cancel_pushButton.setEnabled(False)

    # 샷건 프로젝트
    project_dict = sg.find_one('Project', [['name', 'is', projectName]])

    # 복사하려는 디렉토리 생성
    if not os.path.isdir(pathToCopy):
        os.makedirs(pathToCopy)

    for r in range(tableRow):
        copied = False
        shotCode = self.ui.edit_fileCopy_tableWidget.item(r, self.edit_column_idx_lookup["ShotCode"]).text()

         # 샷코드가 규칙에 맞는지 확인
        if not re.match("s[0-9]+_c[0-9]+", shotCode):
            if not re.match("s[0-9]+_c[a-z]+", shotCode):
                item = QtWidgets.QTableWidgetItem("샷코드가 규칙에 맞지 않습니다.")
                item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
                self.ui.edit_fileCopy_tableWidget.setItem(r, self.edit_column_idx_lookup["Error"], item)

                color = QtGui.QColor(255, 0, 0, 125)
                self.ui.edit_fileCopy_tableWidget.item(r, self.edit_column_idx_lookup["ShotCode"]).setBackground(color)
                QtCore.QCoreApplication.processEvents()
                continue

        # 동일한 샷코드가 있는지 확인
        if shotCode in shotCodeList:
            item = QtWidgets.QTableWidgetItem("이미 존재하는 샷코드입니다.")
            item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
            self.ui.edit_fileCopy_tableWidget.setItem(r, self.edit_column_idx_lookup["Error"], item)

            color = QtGui.QColor(255, 0, 0, 125)
            self.ui.edit_fileCopy_tableWidget.item(r, self.edit_column_idx_lookup["ShotCode"]).setBackground(color)
            QtCore.QCoreApplication.processEvents()
            continue
        shotCodeList.append(shotCode)

        # 해당 프로젝트의 샷코드로 샷건에 edit 태스크인 버전을 가져온다.
        shot_dict = sg.find_one('Shot', [['project', 'is', project_dict], ['code', 'is', shotCode]])
        if shot_dict == None:
            item = QtWidgets.QTableWidgetItem("샷건에 존재하지 않는 샷코드입니다.")
            item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
            self.ui.edit_fileCopy_tableWidget.setItem(r, self.edit_column_idx_lookup["Error"], item)

            color = QtGui.QColor(255, 0, 0, 125)
            self.ui.edit_fileCopy_tableWidget.item(r, self.edit_column_idx_lookup["ShotCode"]).setBackground(color)
            QtCore.QCoreApplication.processEvents()
            continue
        taskName = "edit"
        task_dict = sg.find_one('Task', [['project', 'is', project_dict], ['entity', 'is', shot_dict], ['content', 'is', taskName]])
        if task_dict == None:
            item = QtWidgets.QTableWidgetItem("Edit 태스크가 존재하지 않는 샷코드입니다.")
            item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
            self.ui.edit_fileCopy_tableWidget.setItem(r, self.edit_column_idx_lookup["Error"], item)

            color = QtGui.QColor(255, 0, 0, 125)
            self.ui.edit_fileCopy_tableWidget.item(r, self.edit_column_idx_lookup["ShotCode"]).setBackground(color)
            QtCore.QCoreApplication.processEvents()
            continue
        version_dict = sg.find('Version', [['project', 'is', project_dict], ['entity', 'is', shot_dict], ['sg_task', 'is', task_dict]], ['code', 'sg_path_to_movie'])
        
        # 샷건에 해당 버전 정보가 있는 경우
        if version_dict != []:
            copied = True
            maxValue = int(version_dict[0]["code"].split("v")[1])
            lastVersion = version_dict[0]

            # 가장 최신 버전을 찾는다
            for version in version_dict:
                if maxValue < int(version["code"].split("v")[1]):
                    maxValue = int(version["code"].split("v")[1])
                    lastVersion = version
            
            # 경로가 쓰여져 있는지 확인
            if lastVersion["sg_path_to_movie"] == None:
                item = QtWidgets.QTableWidgetItem("샷건 버전에 경로 정보가 존재하지 않습니다.")
                item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
                self.ui.edit_fileCopy_tableWidget.setItem(r, self.edit_column_idx_lookup["Error"], item)

                color = QtGui.QColor(255, 0, 0, 125)
                self.ui.edit_fileCopy_tableWidget.item(r, self.edit_column_idx_lookup["ShotCode"]).setBackground(color)
                QtCore.QCoreApplication.processEvents()
                continue

            else:
                # 경로가 쓰여있는 경우
                if not os.path.exists(lastVersion["sg_path_to_movie"]):
                    item = QtWidgets.QTableWidgetItem("입력된 경로에 파일이 존재하지 않습니다.")
                    item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
                    self.ui.edit_fileCopy_tableWidget.setItem(r, self.edit_column_idx_lookup["Error"], item)

                    color = QtGui.QColor(255, 0, 0, 125)
                    self.ui.edit_fileCopy_tableWidget.item(r, self.edit_column_idx_lookup["ShotCode"]).setBackground(color)
                    QtCore.QCoreApplication.processEvents()
                    continue

                else:
                    # 해당 경로에 파일이 있는 경우 파일 복사
                    shutil.copy(lastVersion["sg_path_to_movie"], pathToCopy)

                    # 복사된 타입과 버전을 테이블에 적어준다.
                    item = QtWidgets.QTableWidgetItem(lastVersion["code"])
                    item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
                    self.ui.edit_fileCopy_tableWidget.setItem(r, self.edit_column_idx_lookup["Edit"], item)
                    
                    color = QtGui.QColor(0, 255, 0, 125)
                    self.ui.edit_fileCopy_tableWidget.item(r, self.edit_column_idx_lookup["ShotCode"]).setBackground(color)
                    QtCore.QCoreApplication.processEvents()

        else:
            # 샷건에 해당 버전 정보가 없는 경우
            item = QtWidgets.QTableWidgetItem("샷건에 버전 정보가 존재하지 않습니다.")
            item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
            self.ui.edit_fileCopy_tableWidget.setItem(r, self.edit_column_idx_lookup["Error"], item)

            color = QtGui.QColor(255, 0, 0, 125)
            self.ui.edit_fileCopy_tableWidget.item(r, self.edit_column_idx_lookup["ShotCode"]).setBackground(color)
            QtCore.QCoreApplication.processEvents()
            continue

    self.ui.edit_fileCopy_tableWidget.resizeRowsToContents()
    self.ui.edit_fileCopy_tableWidget.resizeColumnsToContents()

    # 버튼을 또 누르는 것을 방지하기 위해서 복사가 완료되면 버튼을 비활성화한다,
    self.ui.edit_copy_pushButton.setEnabled(False)

    # 툴 메뉴 활성화
    self.ui.edit_project_comboBox.setEnabled(True)
    self.ui.edit_excel_radioButton.setEnabled(True)
    self.ui.edit_shotgun_radioButton.setEnabled(True)
    self.ui.edit_pathToCopy_pushButton.setEnabled(True)
    if self.ui.edit_shotgun_radioButton.isChecked():
        self.ui.edit_vendor_comboBox.setEnabled(True)
        self.ui.edit_vendorLoad_pushButton.setEnabled(True)
    if self.ui.edit_excel_radioButton.isChecked():
        self.ui.edit_excelFile_pushButton.setEnabled(True)
        self.ui.edit_excelLoad_pushButton.setEnabled(True)
    self.ui.edit_export_pushButton.setEnabled(True)
    self.ui.edit_cancel_pushButton.setEnabled(True)

    dial = QtWidgets.QMessageBox()
    dial.setWindowTitle("Message")
    dial.setText("편집본 복사가 모두 완료되었습니다.")
    ok_btn = dial.addButton("OK", QtWidgets.QMessageBox.YesRole)
    dial.exec_()

def exportEditExcelFunc(self):
    '''
    테이블의 내용을 엑셀 파일로 
    정리하여 추출하는 함수이다.
    '''
    # 프로젝트 콤보박스 확인
    projectName = self.ui.edit_project_comboBox.currentText()
    if projectName == "":
        dial = QtWidgets.QMessageBox()
        dial.setWindowTitle("Error")
        dial.setText(u"프로젝트를 선택해주세요.")
        ok_btn = dial.addButton("OK", QtWidgets.QMessageBox.YesRole)
        dial.exec_()
        return
    
    # 복사할 경로가 정해졌는지 확인
    pathToCopy = self.ui.edit_pathtocopy_lineEdit.text()
    if pathToCopy == "":
        dial = QtWidgets.QMessageBox()
        dial.setWindowTitle("Error")
        dial.setText(u"복사할 경로를 지정해주세요.")
        ok_btn = dial.addButton("OK", QtWidgets.QMessageBox.YesRole)
        dial.exec_()
        return
    
    # 엑셀 라디오 버튼이 체크되어있으면 엑셀 파일 경로가 정해졌는지 확인
    if self.ui.edit_excel_radioButton.isChecked():
        excelPath = self.ui.edit_excelFile_lineEdit.text()
        if excelPath == "":
            dial = QtWidgets.QMessageBox()
            dial.setWindowTitle("Error")
            dial.setText(u"엑셀 파일을 먼저 선택해주세요.")
            ok_btn = dial.addButton("OK", QtWidgets.QMessageBox.YesRole)
            dial.exec_()
            return
    
    # 벤더 라디오 버튼이 체크되어있으면 벤더 콤보박스가 체크되었는지 확인
    if self.ui.edit_shotgun_radioButton.isChecked():
        vendor = self.ui.edit_vendor_comboBox.currentText()
        if vendor == "":
            dial = QtWidgets.QMessageBox()
            dial.setWindowTitle("Error")
            dial.setText(u"벤더를 먼저 선택해주세요.")
            ok_btn = dial.addButton("OK", QtWidgets.QMessageBox.YesRole)
            dial.exec_()
            return

    # 샷코드를 읽어왔는지 확인
    tableRow = self.ui.edit_fileCopy_tableWidget.rowCount()
    if tableRow == 0:
        dial = QtWidgets.QMessageBox()
        dial.setWindowTitle("Error")
        dial.setText("Copy할 데이터가 존재하지 않습니다.")
        ok_btn = dial.addButton("OK", QtWidgets.QMessageBox.YesRole)
        dial.exec_()
        return
    
    # 엑셀 만들기
    excelFileName = "{}_edit_copy.xlsx".format(projectName)
    excelFilePath = os.path.join(pathToCopy, excelFileName)

    write_wb = openpyxl.Workbook()
    write_ws = write_wb.create_sheet('edit_copy')
    write_ws = write_wb.active

    # 엑셀 제목
    write_ws.cell(1, 1, "ShotCode")
    write_ws.cell(1, 2, "Edit")
    write_ws.cell(1, 3, "Error")

    # 엑셀 내용 입력
    for r in range(tableRow):
        shotCode = self.ui.edit_fileCopy_tableWidget.item(r, self.edit_column_idx_lookup["ShotCode"]).text()
        write_ws.cell(r+2, 1, shotCode)

        if self.ui.edit_fileCopy_tableWidget.item(r, self.edit_column_idx_lookup["Edit"]):
            edit = self.ui.edit_fileCopy_tableWidget.item(r, self.edit_column_idx_lookup["Edit"]).text()
            write_ws.cell(r+2, 2, edit)
        
        if self.ui.edit_fileCopy_tableWidget.item(r, self.edit_column_idx_lookup["Error"]):
            error = self.ui.edit_fileCopy_tableWidget.item(r, self.edit_column_idx_lookup["Error"]).text()
            write_ws.cell(r+2, 3, error)

    # 엑셀 파일 셀 너비 및 높이 지정
    for col in write_ws.columns:
        max_width = 0
        col_name = col[0].column_letter
        for cell in col:
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            try:
                if len(str(cell.value.encode('utf-8'))) > max_width:
                    max_width = len(str(cell.value.encode('utf-8')))
            except:
                continue
        adjusted_width = (max_width + 2) * 1.2
        write_ws.column_dimensions[col_name].width = adjusted_width
    
    for i, row in enumerate(write_ws):
        default_height = 15
        write_ws.row_dimensions[i+1].height = default_height
    
    # 엑셀 파일 저장
    write_wb.save(excelFilePath)

    dial = QtWidgets.QMessageBox()
    dial.setWindowTitle("Message")
    dial.setText("해당 경로로 엑셀 파일이 저장되었습니다.")
    ok_btn = dial.addButton("OK", QtWidgets.QMessageBox.YesRole)
    dial.exec_()